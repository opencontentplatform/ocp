"""Generate reports from definitions.

This uses the openpyxl library (https://openpyxl.readthedocs.io/en/stable),
which is a Python library to read/write Excel 2010 xlsx/xlsm files.

So to use this job, you need to 'pip install <module>' for these:
  openpyxl    <-- core library
  lxml        <-- suggested by openpyxl; used when creating large files
  defusedxml  <-- suggested by openpyxl for security; protects against
                  quadratic blowup and billion laughs xml attacks

Functions:
  startJob : standard job entry point
  getNodesWithoutDomains : get the first set of nodes
  getNodesWithDomains : get the second set of nodes
  processResults : do the work

Author: Chris Satterthwaite (CS)
Contributors:
Version info:
  1.0 : (CS) Created Jan 3, 2018

"""
import sys
import traceback
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, GradientFill

## From openContentPlatform
from utilities import verifyJobRuntimePath, getApiQueryResultsFull


def recurseResult(runtime, worksheet, result, col, row, header1, header2):
	"""Recurse through the results, creating the desired row."""
	objectId = result.get('identifier')
	objectType = result.get('class_name')
	objectLabel = result.get('label')
	data = result.get('data', {})
	startCol = col
	## Set RemoveEmptyAttributes=False to retain order with empty attributes
	for key,value in data.items():
		if isinstance(value, (list, set, tuple, dict)):
			value = str(value)
		worksheet.cell(column=col, row=row, value=value)
		if len(header2) < col:
			header2.append(key)
		col += 1
	header1[objectLabel] = {'start_row':1, 'start_column':startCol, 'end_row':1, 'end_column':col-1}

	children = result.get('children', {})
	if len(children) > 0:
		flag = False
		for childResult in children:
			if flag:
				row += 1
				parentSegment = []
				## Now copy the previous parent segment to the start of this row
				for x in range(1, col+1):
					worksheet.cell(column=x, row=row, value=worksheet.cell(row = row-1, column = x).value)
			row = recurseResult(runtime, worksheet, childResult, col, row, header1, header2)
			flag = True
	return row


def transformResults(runtime, queryResults, worksheet):
	"""Convert the API results into the desired reporting layout."""
	## Since ResultsFormat of the result is Nested-Simple, each of the
	## linchpin results is an independent topology result (for a Excel row)
	row = 1
	header1 = {}
	header2 = []
	for result in queryResults:
		row = recurseResult(runtime, worksheet, result, 1, row, header1, header2)
		row += 1

	## Now insert the header rows
	worksheet.insert_rows(idx=1, amount=2)

	## First row for object types (column groups), will have a fill every other
	## Object. So if there are 3 object types, the 1st and 3rd will be filled.
	## Intent is to make object types stand out naturally to the visible eye.
	header1Font = Font(b=True, color='4682B4')
	header1Fill = PatternFill('solid', fgColor='F2F3F3')
	fillFlag = True
	for label,mergeIndex in header1.items():
		worksheet.merge_cells(**mergeIndex)
		c = worksheet.cell(column=mergeIndex['start_column'], row=1)
		if fillFlag:
			c.fill = header1Fill
		fillFlag = not fillFlag
		c.font = header1Font
		c.value = label

	header2Font = Font(color='4682B4')
	header2Fill = GradientFill(stop=('F2F3F3', 'FFFFFF'), degree=90)
	## Second row for attribute names (each column) will have a gradient fill
	## of light gray, top down.
	for x in range(1, len(header2)+1):
		c = worksheet.cell(column=x, row=2)
		c.fill = header2Fill
		c.font = header2Font
		c.value = header2[x-1]

	## Now freeze the header rows, above the cell named in the freeze call
	worksheet.freeze_panes = worksheet.cell(column=1, row=3)

	## end transformResults
	return


def startJob(runtime):
	"""Standard job entry point.

	Arguments:
	  runtime (dict)   : object used for providing input into jobs and tracking
	                     the job thread through the life of its runtime.
	"""
	try:
		## Establish our runtime working directory
		jobRuntimePath = verifyJobRuntimePath(__file__)
		runtime.logger.report('path jobRuntimePath: {jobRuntimePath!r}', jobRuntimePath=jobRuntimePath)

		## Initialize a workbook
		wb = Workbook()
		worksheets = runtime.parameters.get('worksheets')
		firstSheet = True
		## Go through each worksheet definition set in the job
		for definition in worksheets:
			try:
				queryName = definition.get('queryName')
				sheetName = definition.get('sheetName')
				worksheet = wb.active
				if firstSheet:
					worksheet.title = sheetName
					firstSheet = False
				else:
					wb.create_sheet(sheetName)
					worksheet = wb.get_sheet_by_name(sheetName)

				## Read the query definition
				queryFile = os.path.join(runtime.env.universalJobPkgPath, 'reportWorkbook', 'input', queryName + '.json')
				if not os.path.isfile(queryFile):
					raise EnvironmentError('JSON query file does not exist: {}'.format(queryFile))
				queryContent = None
				with open(queryFile) as fp:
					queryContent = json.load(fp)

				runtime.logger.report(' Requesting queryName {queryName!r}', queryName=queryName)
				queryResults = getApiQueryResultsFull(runtime, queryContent, resultsFormat='Nested-Simple', headers={'removeEmptyAttributes': False})
				if queryResults is None or len(queryResults) <= 0:
					runtime.logger.debug('No results found for queryName {queryName!r}', queryName=queryName)
					continue

				## Special case since we need Nested format; if query was only
				## looking for a single object, the nested format drops down to
				## a Flat format, with a list of objects.
				if 'objects' in queryResults:
					queryResults = queryResults['objects']

				## Convert the API results into a desired layout for our needs here
				transformResults(runtime, queryResults, worksheet)

			except:
				stacktrace = traceback.format_exception(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2])
				runtime.logger.error('Failure parsing queryName {queryName!r}: {stacktrace!r}', queryName=queryName, stacktrace=stacktrace)

		## Save the workbook as a file
		reportName = '{}.xlsx'.format(runtime.parameters.get('reportFile'))
		reportFile = os.path.join(jobRuntimePath, reportName)
		wb.save(reportFile)

		## Update the runtime status to success
		if runtime.getStatus() == 'UNKNOWN':
			runtime.status(1)

	except:
		runtime.setError(__name__)

	## end startJob
	return
